# pip install openpyxl

import openpyxl as excel
from openpyxl import chart

work_book = excel.load_workbook('./data/transaction.xlsx')
sheet = work_book["transaction"]
num_rows = sheet.max_row
for i_row in range(2, num_rows + 1):
    cell = sheet.cell(i_row, 3)
    reduce = cell.value * 0.9
    reduced_cell = sheet.cell(i_row, 4)
    reduced_cell.value = reduce

ref_values = chart.Reference(sheet, 
                      min_row=2, 
                      max_row=sheet.max_row, 
                      min_col=4, 
                      max_col=4)

bar = chart.BarChart()
bar.add_data(ref_values)
sheet.add_chart(bar, "E2")

work_book.save("./data/transaction.xlsx")
